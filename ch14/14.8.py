import openpyxl, csv, os

for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excel_file)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        csv_file = open(excel_file[:-5] + '_' + sheet_name + '.csv', 'w', newline='')
        csv_writer = csv.writer(csv_file)

        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row_num, column=col_num).value)
            csv_writer.writerow(row_data)
        csv_file.close()

