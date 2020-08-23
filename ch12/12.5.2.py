import openpyxl

wb = openpyxl.Workbook()
print("创建sheet前: ", end='')
print(wb.sheetnames)

wb.create_sheet()
print("创建sheet后: ", end='')
print(wb.sheetnames)

wb.create_sheet(index=0, title='First Sheet')
print("在0号位置创建First Sheet后: ", end='')
print(wb.sheetnames)

wb.create_sheet(index=2, title='Middle Sheet')
print("在2号位置创建First Sheet后: ", end='')
print(wb.sheetnames)

wb.remove(wb['Middle Sheet'])
print("移除Middle Sheet后: ", end='')
print(wb.sheetnames)

wb.remove(wb['Sheet1'])
print("移除Sheet1后: ", end='')
print(wb.sheetnames)