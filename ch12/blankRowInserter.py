import openpyxl, sys

row_index = sys.argv[1]
row_space = sys.argv[2]
xlsx_file_path =sys.argv[3]

wb = openpyxl.load_workbook('example_copy.xlsx')
sheet = wb.active

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(max_row, 0, -1):
    for j in range(1, max_col+1):
        if i >= (row_index):
            sheet.cell(row=i+row_space,column=j).value = sheet.cell(row=i,column=j).value
        if i < (row_index + row_space) and i >= row_index:
            sheet.cell(row=i,column=j).value = ' '

wb.save(xlsx_file_path)
        

