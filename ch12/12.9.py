import openpyxl

wb =  openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')


wb_formulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wb_formulas.active
print(sheet['A3'].value)
wb.save('writeFormulaCopy.xlsx')


# wb_data_only = openpyxl.load_workbook('writeFormulaCopy.xlsx', data_only=True)
# sheet = wb_data_only.active
# # print(sheet['A3'].value)
# wb.save('writeFormulaCopy.xlsx')
# print(sheet['A3'].value)

