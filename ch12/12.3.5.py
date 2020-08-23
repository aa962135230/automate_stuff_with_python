import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

print(tuple(sheet['A1': 'C3']))

for row_of_cell_objects in sheet['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print("--- END OF ROW ---")


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
print(list(sheet.columns)[1])

for cell_obj in list(sheet.columns)[1]:
    print(cell_obj.value)

    
i = sheet.max_row
j = sheet.max_column

for row in range(i):
    for col in range(j):
        if col + 1 == j:
            print(sheet.cell(row+1, col+1).value)
        else:
            print(sheet.cell(row+1, col+1).value, ' ', end='')





