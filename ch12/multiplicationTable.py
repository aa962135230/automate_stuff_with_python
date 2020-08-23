import openpyxl, sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

num = int(sys.argv[1])
sheet.freeze_panes = 'B2'

for i in range(1, num + 1):
    for j in range(1, num + 1):
        if i == 1 and j >= 1:
            #第一行
            sheet.cell(row=1, column=j+1).font = Font(bold=True)
            sheet.cell(row=1, column=j+1).value = j

        if i >= 1 and j == 1:
            #第一列
            sheet.cell(row=i+1, column=1).font = Font(bold=True)
            sheet.cell(row=i+1, column=1).value = i
         
        
        #乘法表    
        sheet.cell(row=i+1,column=j+1).value = i*j

wb.save('multiplication.xlsx')

        